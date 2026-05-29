"""
extract-budget.py
Reads "BUDGET (1).xlsx" and writes src/data/budget.json

Run from project root:
    python3 scripts/extract-budget.py

Install dep if needed:
    pip install openpyxl
"""

import json, os
from pathlib import Path
import openpyxl

ROOT = Path(__file__).parent.parent
wb = openpyxl.load_workbook(ROOT / "BUDGET (1).xlsx", data_only=True)

# Use the most recent month sheet (skip CARD USED - LIST)
month_sheets = [s for s in wb.sheetnames if s != "CARD USED - LIST"]
sheet_name = month_sheets[-1]
ws = wb[sheet_name]

rows = list(ws.iter_rows(values_only=True))

# ── Income ──────────────────────────────────────────────────────────────────
total_income = 0
raw_free_spend = 0

for row in rows:
    if row[0] == "Total" and isinstance(row[1], (int, float)):
        total_income = float(row[1])
    if isinstance(row[0], str) and "raw free spend" in row[0].lower():
        raw_free_spend = float(row[1]) if isinstance(row[1], (int, float)) else 0

# ── Expenses ─────────────────────────────────────────────────────────────────
FIXED_LABELS = {
    "Rent", "Roth IRA", "Subscriptions", "Phone Bill", "Gym Membership",
    "Internet", "MTA", "Savings", "Seperate Account",
}
VARIABLE_LABELS = {
    "Groceries", "Eating out", "Coffee/ drinks", "Ubers",
    "Fun", "Travel Fund", "Emergency Savings", "Investing", "misc", "claude",
}

fixed_expenses = []
variable_expenses = []
in_fixed = False
in_variable = False

for row in rows:
    if row[0] == "Fixed Expenses":
        in_fixed, in_variable = True, False
        continue
    if row[0] == "Variable Expenses":
        in_fixed, in_variable = False, True
        continue
    if in_fixed and row[0] in FIXED_LABELS:
        fixed_expenses.append({
            "label": row[0],
            "budget": float(row[1] or 0),
            "actual": float(row[2] or 0),
        })
    if in_variable and row[0] in VARIABLE_LABELS:
        variable_expenses.append({
            "label": row[0],
            "budget": float(row[1] or 0),
            "actual": float(row[2] or 0),
        })

# ── Savings ──────────────────────────────────────────────────────────────────
savings_row = next((r for r in fixed_expenses if r["label"] == "Savings"), None)
separate_row = next((r for r in fixed_expenses if r["label"] == "Seperate Account"), None)
savings_target = (savings_row["budget"] if savings_row else 0) + (separate_row["budget"] if separate_row else 0)
savings_actual = (savings_row["actual"] if savings_row else 0) + (separate_row["actual"] if separate_row else 0)

# ── Free spend ───────────────────────────────────────────────────────────────
variable_actual_spent = sum(r["actual"] for r in variable_expenses)
free_spend_remaining = raw_free_spend - variable_actual_spent

output = {
    "month": sheet_name,
    "totalIncome": total_income,
    "rawFreeSpend": raw_free_spend,
    "freeSpendRemaining": free_spend_remaining,
    "variableActualSpent": variable_actual_spent,
    "savingsTarget": savings_target,
    "savingsActual": savings_actual,
    "fixedExpenses": fixed_expenses,
    "variableExpenses": variable_expenses,
}

out_path = ROOT / "src" / "data" / "budget.json"
out_path.parent.mkdir(parents=True, exist_ok=True)
out_path.write_text(json.dumps(output, indent=2))

print(f"✅  Budget extracted from '{sheet_name}' sheet")
print(f"    Income:        ${total_income:,.2f}")
print(f"    Free to spend: ${free_spend_remaining:,.2f} / ${raw_free_spend:,.2f}")
print(f"    Savings:       ${savings_actual:,.2f} / ${savings_target:,.2f}")
print(f"    Written to:    src/data/budget.json")
